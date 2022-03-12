from robot.libraries.BuiltIn import BuiltIn
import time

def fncGetValues(loc, sheetName, testCase):
    # import openpyxl module to use excel operations
    import openpyxl
    from openpyxl import load_workbook

    # Create workbook object and capture the active sheet
    wb = load_workbook(loc, data_only=True)
    sheet = wb.active

    # Designate the actual sheetname
    sheet = wb[sheetName]

    # Leaving the header row, go through each line to find a match with test case you want to execute
    for i in range(2, sheet.max_row + 1):
        if (sheet.cell(row=i, column=1).value == testCase):

            # if the matching test case is found, create a dictionary with data in second column
            getValues = {sheet.cell(row=1, column=2).value: sheet.cell(row=i, column=2).value}

            # Once the data dictionary is added, keep on adding rest of the items in that row till the end of columns
            for j in range(3, sheet.max_column + 1):
                getValues.update({sheet.cell(row=1, column=j).value: sheet.cell(row=i, column=j).value})

            break

    # Return the dictionary
    return getValues


def fncSelectDDElement(sl, ddelement, uValue):
    # if the selected value is same as value being sent, do not do anything
    if uValue == sl.get_text(ddelement):
        BuiltIn().log('The value is already selected')

    # if the value is sent and not same as selected item, proceed with selecting the new value from Dropdown element
    elif uValue != '':
        ddlitem = 'xpath://li[@class="rddlItem" and text()="' + uValue + '"]'
        sl.click_element(ddelement)
        time.sleep(2)
        sl.click_element(ddlitem)

    # if the value is empty, select the emtry record on top of the list
    else:
        ddlitem = 'xpath://li[@class="rddlItem"]'
        sl.click_element(ddelement)
        time.sleep(2)
        sl.click_element(ddlitem)
