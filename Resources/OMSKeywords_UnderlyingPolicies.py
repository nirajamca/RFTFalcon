from robot.libraries.BuiltIn import BuiltIn
from robot.api.deco import keyword
import AccessRepository as AR
import CommonFunctions as CF
import openpyxl
import platform


def fncAddUnderLyingPolicyCoverage(sl, uLayerNumber, uPolicyType, uCarrier, uULCoverageCode, uULCoverageType,
                                   uPremium, uULRatedSUPolicy, uRCCLimit, uRetention):

    # Click Add button to add policies
    sl.click_button(AR.idbtnULPolicyAdd)
    BuiltIn().sleep(2)

    # Layer
    sl.input_text(AR.idtxtULPoliciesLayerNumber, int(uLayerNumber))

    # Policy Type
    CF.fncSelectDDElement(sl, AR.idddULPPolicyType, uPolicyType)

    # Carrier
    sl.input_text(AR.idtxtULCarrier, str(uCarrier))
    BuiltIn().sleep(1)
    sl.press_keys(AR.idtxtULCarrier, 'ARROW_DOWN')
    BuiltIn().sleep(1)
    sl.press_keys(AR.idtxtULCarrier, 'RETURN')


    # Under Lying Coverage Code
    CF.fncSelectDDElement(sl, AR.idddULCoverageCode, uULCoverageCode)

    # Coverage Type
    CF.fncSelectDDElement(sl, AR.idddULCoverageType, uULCoverageType)

    # Premium
    sl.input_text(AR.idtxtULPoliciesPremium, uPremium)

    # Rated Scheduled Underlying Policy
    CF.fncSelectDDElement(sl, AR.idddULRatedSUPolicy, uULRatedSUPolicy)

    # Coverage table
    # General Aggregate Checkbox
    sl.click_element(AR.idchkULRiskCoverageControl)

    # Limit
    sl.input_text(AR.idtxtULRCCLimit, uRCCLimit)

    # Deductible/Retention
    sl.input_text(AR.idtxtULRCCRetension, uRetention)

    # Save the coverage
    sl.click_button(AR.idbtnULRCCSave)
    BuiltIn().sleep(2)



@keyword
def Go_To_UL_Policies():
    sl = BuiltIn().get_library_instance('SeleniumLibrary')
    try:
        sl.click_element(AR.UnderlyingPoliciesTab)

    except Exception as e1:
        BuiltIn().log('Not able to find the button')

@keyword
def Delete_Existing_Policies_if_Required():
    uCompany = BuiltIn().get_variable_value('${Company}')
    myData = CF.getData(uCompany, 'UnderlyingPolicies')
    uDeletePolicies = myData['Option1']

    if uDeletePolicies == 'Yes':
        sl = BuiltIn().get_library_instance('SeleniumLibrary')
        try:
            sl.click_button(AR.idbtnULPDeleteAll)
            BuiltIn().sleep(2)
            sl.click_button(AR.idbtnULPConfirmYes)

        except Exception as e1:
            BuiltIn().log('There are no underlying policies to delete')


@keyword
def Add_Layers():
    uCompany = BuiltIn().get_variable_value('${Company}')
    myData = CF.getData(uCompany, 'UnderlyingPolicies')
    uToAddLayers = myData['Option2']
    uLayers = myData['Option3']

    if uToAddLayers == 'Yes':
        if platform.system() == 'Windows':
            loc = 'Data\\FalconTestData.xlsx'
        elif platform.system() == 'Darwin':
            loc = 'Data/FalconTestData.xlsx'
        wb = openpyxl.load_workbook(loc)
        sheet = wb['UnderlyingPolicies']

        for i in range(1, int(uLayers)+1):
            uLayerNumber = sheet.cell(row=i, column=1).value
            uPolicyType = sheet.cell(row=i, column=2).value
            uCarrier = sheet.cell(row=i, column=3).value
            uULCoverageCode = sheet.cell(row=i, column=4).value
            uULCoverageType = sheet.cell(row=i, column=5).value
            uPremium = sheet.cell(row=i, column=6).value
            uULRatedSUPolicy = sheet.cell(row=i, column=7).value
            uRCCLimit = sheet.cell(row=i, column=8).value
            uRetention = sheet.cell(row=i, column=9).value

            sl = BuiltIn().get_library_instance('SeleniumLibrary')
            fncAddUnderLyingPolicyCoverage(sl, uLayerNumber, uPolicyType, uCarrier, uULCoverageCode,
                                           uULCoverageType, uPremium, uULRatedSUPolicy, uRCCLimit, uRetention)



@keyword
def Save_Layers():
    sl = BuiltIn().get_library_instance('SeleniumLibrary')
    # Click Next to save UW Questions
    sl.click_button(AR.idbtnOneShieldNext)
    BuiltIn().sleep(2)

    uCompany = BuiltIn().get_variable_value('${Company}')
    myData = CF.getData(uCompany, 'UnderlyingPolicies')
    uToAddLayers = myData['Option2']

    if uToAddLayers == 'Yes':
        # Verify if the record has been successfully created
        sl.element_should_contain(AR.idSubmissionBanner, 'Record has been saved')



