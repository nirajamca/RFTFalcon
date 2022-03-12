from robot.libraries.BuiltIn import BuiltIn
from robot.api.deco import keyword
from robot.libraries.OperatingSystem import OperatingSystem
import AccessRepository as AR
import CommonFunctions as CF

# def getData():
#     uEnvironment = BuiltIn().get_variable_value('${Env}')
#     loc = 'Data\\' + uEnvironment + '.xlsx'
#     SheetName = BuiltIn().get_variable_value('${Company}')
#     testcase = 'UWQuestions'
#     return CF.fncGetValues(loc, SheetName, testcase)

def fncGetFinalPremiumFromRater(uHref):
    from openpyxl import load_workbook
    from robot.libraries.OperatingSystem import OperatingSystem
    import locale

    uRaterFile = uHref.split('=')[1]

    myDownloads = OperatingSystem().get_environment_variable('USERPROFILE')
    uRaterFileLocal = myDownloads + '\\Downloads\\' + uRaterFile

    # Create workbook object and capture the active sheet
    wb = load_workbook(uRaterFileLocal, data_only=True)
    sheet = wb.active

    # Designate the actual sheetname
    sheet = wb['XS Final Premium Calculation']

    # Get Final Adjustment Premium
    uFinalPremium = sheet.cell(row=41, column=2).value

    ExFinalPremium = locale.currency(uFinalPremium, grouping = True)

    return ExFinalPremium

@keyword
def Verify_Premium_with_Rater():
    sl = BuiltIn().get_library_instance('SeleniumLibrary')
    uHref = sl.get_element_attribute(AR.xPathlnkQRaterFile, 'HREF')
    BuiltIn().log(uHref)

    sl.click_element(AR.xPathlnkQRaterFile)
    BuiltIn().sleep(3)

    BuiltIn().log(sl.get_text(AR.idlblQPremium))
    BuiltIn().log(fncGetFinalPremiumFromRater(uHref))














